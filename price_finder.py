from openpyxl import load_workbook

MENU = '''Welcome:
check price: -c
change workbook: -wb
quit: -q'''

CHECK_PRICE_PROMPT = 'Enter a customer or enter -q to quit'

# Define column letters for various data in the Excel sheet
CUSTOMER_COLUMN = 'A'
PRODUCT_COLUMN = 'C'
PRICE_COLUMN = 'F'
QTY_COLUMN = 'E'
UNIT_COLUMN = 'D'
DATE_COLUMN = 'B'

SORTED_NUMBER = 3  # Number of prices users want to see

workbook_name = 'test.xlsx'  # Default workbook name
workbook = load_workbook(workbook_name)  # Load the workbook
sheet = workbook.active  # Set the active sheet of the workbook


def print_current_workbook():
    '''
    Function print_current_workbook
    print which workbook user is using
    '''
    print(f"Currently open workbook: {workbook_name}")


def clean_up(price_list: list) -> list:
    '''
    Function clean_up
    Sorts a list of prices and retains the last 'SORTED_NUMBER' elements.
    Args:
        price_list(list): The list of prices to be sorted.
    Returns:
        sorted_list(list): The sorted list containing the most recent 'SORTED_NUMBER' prices.
    '''
    sorted_list = sorted(price_list)
    if len(sorted_list) >= SORTED_NUMBER:
        index = len(sorted_list) - SORTED_NUMBER
        sorted_list = sorted_list[index:]
        return sorted_list
    if len(sorted_list) < SORTED_NUMBER:
        return sorted_list


def get_info(product_name: str, row: str, product: str) -> str:
    '''
    Function get_info:
    Retrieves information about a product.
    Args:
        product_name(str): The name of the product.
        row(str): The row where the product information is located.
        product(str): The product information.
    Returns:
        info(str): Information about the product.
    '''
    # Gather product information from respective columns in the Excel sheet
    customer = str(sheet[CUSTOMER_COLUMN + row].value).lower().strip()
    price = str(sheet[PRICE_COLUMN + row].value).lower().strip()
    qty = str(sheet[QTY_COLUMN + row].value).lower().strip()
    unit = str(sheet[UNIT_COLUMN + row].value).lower().strip()
    date = str(sheet[DATE_COLUMN + row].value).lower().strip()
    info = str(f'{date}: {customer} purchased {product} with {qty} {unit}, each {price}')
    return info


def check_customer(customer_name):
    '''
    Function check_customer
    Checks if a customer exists in the Excel sheet.
    Args:
        customer_name(str): The name of the customer to check.
    Returns:
        customer_found(bool): True if the customer is found, False otherwise.
    '''
    customer_found = False
    for column in sheet[CUSTOMER_COLUMN]:  # Loop column A to get all the customer
        customer = str(column.value).lower().strip()
        if customer_name in customer:  # Loop customer's names to see find the rows customer_name is at
            customer_found = True  # True if the customer is found
    return customer_found


def find_price_for_all_products(product_name: str) -> list:
    '''
    Function find_price_for_all_products
    Finds prices for a specific product across all customers.
    Args:
        product_name(str): The name of the product.
    Returns:
        price_list(list): List containing product info for all customers.
    '''
    price_list = []
    price_found = False

    for products in sheet[PRODUCT_COLUMN]:  # Check each product in the column
        product = str(products.value).lower().strip()
        product_row = str(products.row)  # Use the rows to find product

        if product_name in product:
            price_found = True  # Set price_found to True if the product is found
            info = get_info(product_name, product_row, product)  # Get information about the product
            price_list.append(info)

    if not price_found and product_name != 'done':  # Let the user know if price is not found
        print(f'{product_name} price not found')

    return price_list


def find_price_for_customer(customer_name: str, product_name: str) -> list:
    '''
    Function find_price_for_customer
    Finds prices for a specific product for a given customer.
    Args:
        customer_name(str): The name of the customer.
        product_name(str): The name of the product.
    Returns:
        price_list(list): List containing product info for the customer.
    '''
    price_list = []
    price_found = False

    for column in sheet[CUSTOMER_COLUMN]:  # Check each customer in the column
        customer = str(column.value).lower().strip()

        if customer_name in customer:  # Check customer's names to see find the rows customer_name is at
            customer_row = str(column.row)  # Use the rows to find product
            product = str(sheet[PRODUCT_COLUMN + customer_row].value).lower().strip()

            if product_name in product:
                price_found = True  # Set price_found to True if the product is found
                info = get_info(product_name, customer_row, product)  # Get information about the product
                price_list.append(info)

    if not price_found and product_name != 'done':  # Let the user know if price is not found
        print(f'{product_name} price not found')

    return price_list


def find_price(customer_name: str, product_names: list) -> None:
    '''
    Function find_price
    Finds prices for specific products and customers.
    Args:
        customer_name(str): The name of the customer.
        product_names(list): List of product names.
    Returns:
        None
    '''
    for product_name in product_names:  # Loop all the products
        price_list = []  # Save all the information and print later

        if customer_name == '-a':  # Get info from all the products if -a was enter
            price_list.extend(find_price_for_all_products(product_name))
        else:  # Get info according to customer's name
            price_list.extend(find_price_for_customer(customer_name, product_name))

        cleaned_list = clean_up(price_list)  # Sort the list using clean_up function
        for item in cleaned_list:
            print(item)


def run():
    '''
    Function run
    Handles the main program flow for interacting with customers and products.
    Args: None
    Returns: None
    '''
    print(CHECK_PRICE_PROMPT)
    products = []  # All the product that user want to check
    c_name = input('Customer: ')
    while c_name != '-q':
        found = check_customer(c_name)
        if found is False and c_name != '-a':  # Check if the name inputted is in the sheet
            print(f'customer {c_name} not found')
            run()
        p_name = ''
        while p_name != 'done':
            p_name = str(input('Product or enter done to end: '))
            products.append(p_name)
        find_price(c_name, products)
        products.clear()
        print(CHECK_PRICE_PROMPT)
        c_name = input('Customer: ')  # Repeat until user enter -q
    print('\n')
    return main()


def load_excel_data(file_path):
    '''
    Function load_excel_data
    Activates an Excel file from the specified file path.
    Args:
        file_path(str): The file path of the Excel workbook to be loaded.
    Returns:
        file active: The active worksheet of the loaded Excel workbook.
        None: If the specified file is not found or cannot be loaded.
    '''
    try:
        workbook = load_workbook(file_path)
        return workbook.active
    except:
        print("\nFile not found or cannot be loaded.\n")
        return None


def main():
    print_current_workbook()  # Let user know which Excel file is open
    print(MENU)
    command = input('What would you like to do? ')
    command = command.lower()
    if command == '-c':
        run()
    elif command == '-wb':
        new_workbook_path = input("Enter the path for the new workbook file: ")
        new_sheet = load_excel_data(new_workbook_path)
        if new_sheet:  # Change to another Excel file
            global workbook, sheet, workbook_name
            workbook = load_workbook(new_workbook_path)
            sheet = workbook.active
            workbook_name = new_workbook_path
            print("\nWorkbook changed successfully!\n")
        main()  # Restart the main menu
    elif command == '-q':
        workbook.close()  # Close the workbook when exiting
        exit()
    else:
        print("Invalid command. Please choose a valid option.\n")
        main()


if __name__ == "__main__":
    main()
